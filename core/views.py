import csv
import io
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib import messages
from django.db import IntegrityError
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from django.views.decorators.http import require_http_methods

from .forms import AttendanceImportForm, AttendanceRecordForm
from .models import AttendanceRecord

# ===== Auth UI pages =====
def login_ui(request):
    return render(request, "auth/login.html")

def signup_ui(request):
    return render(request, "auth/signup.html")


# ===== Admin Dashboard UI pages =====
def admin_dashboard(request):
    return render(request, "admin/dashboard.html", {"current": "dashboard"})

def admin_analytics(request):
    return render(request, "admin/analytics.html", {"current": "analytics"})

def admin_employee_management(request):
    return render(request, "admin/employee_management.html", {"current": "employees"})

def admin_leave_approval(request):
    return render(request, "admin/leave_approval.html", {"current": "leave"})

def admin_payroll(request):
    return render(request, "admin/payroll.html", {"current": "payroll"})

def admin_reports(request):
    return render(request, "admin/reports.html", {"current": "reports"})

def admin_shift_scheduling(request):
    return render(request, "admin/shift_scheduling.html", {"current": "scheduling"})

def admin_system_administration(request):
    return render(request, "admin/system_administration.html", {"current": "system"})


# ===== CRUD Helper Functions =====
def _normalize_event_type(value: str) -> str:
    """Normalize various event type strings to IN/OUT/UNKNOWN."""
    if not value:
        return AttendanceRecord.EVENT_UNKNOWN
    v = value.strip().upper()
    
    in_variations = ["IN", "CHECK IN", "TIME IN", "CLOCK IN", "ENTRY", "ENTRY_TIME"]
    out_variations = ["OUT", "CHECK OUT", "TIME OUT", "CLOCK OUT", "EXIT", "EXIT_TIME"]
    
    if v in in_variations:
        return AttendanceRecord.EVENT_IN
    if v in out_variations:
        return AttendanceRecord.EVENT_OUT
    return AttendanceRecord.EVENT_UNKNOWN


def _parse_timestamp(value: str):
    """Parse timestamp from various common formats."""
    if not value:
        return None

    value = str(value).strip()

    # Try Django's built-in parser first
    dt = parse_datetime(value)
    if dt:
        return dt

    # Try common datetime formats
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    return None


def _read_csv_file(file_obj):
    """Read and parse CSV file, handling encoding issues."""
    raw = file_obj.read()
    
    # Try UTF-8 first (with BOM stripping)
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"]
    text = None
    
    for encoding in encodings:
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    
    if text is None:
        raise ValueError("Could not decode file with any supported encoding")
    
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        # Normalize header names (trim spaces)
        normalized_row = {}
        for k, v in row.items():
            key = k.strip() if k else k
            val = v.strip() if isinstance(v, str) else v
            normalized_row[key] = val
        rows.append(normalized_row)
    
    return rows


def _read_excel_file(file_obj):
    """Read and parse Excel file (.xls or .xlsx)."""
    try:
        import openpyxl
        use_openpyxl = True
    except ImportError:
        use_openpyxl = False
    
    try:
        import pandas as pd
        use_pandas = True
    except ImportError:
        use_pandas = False
    
    if not use_openpyxl and not use_pandas:
        raise ImportError("No Excel library available. Install openpyxl or pandas.")
    
    # Try pandas first (more flexible)
    if use_pandas:
        try:
            df = pd.read_excel(file_obj, sheet_name=0)
            # Convert DataFrame to list of dicts
            rows = df.fillna('').to_dict('records')
            # Normalize headers
            normalized = []
            for row in rows:
                new_row = {}
                for k, v in row.items():
                    key = str(k).strip() if k else k
                    val = str(v).strip() if v else ''
                    new_row[key] = val
                normalized.append(new_row)
            return normalized
        except Exception as e:
            raise ValueError(f"Error reading Excel with pandas: {e}")
    
    # Fall back to openpyxl
    if use_openpyxl:
        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
            
            # Read header
            headers = []
            for cell in ws[1]:
                headers.append(cell.value or '')
            
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, header in enumerate(headers):
                    key = str(header).strip() if header else ''
                    val = row[i] if i < len(row) else None
                    val_str = str(val).strip() if val else ''
                    row_dict[key] = val_str
                if any(row_dict.values()):  # Skip completely empty rows
                    rows.append(row_dict)
            
            return rows
        except Exception as e:
            raise ValueError(f"Error reading Excel with openpyxl: {e}")


def _map_csv_row(row: dict, device_source: str = "") -> dict:
    """
    Map CSV columns to AttendanceRecord fields.
    Tries multiple column name variations for flexibility.
    """
    # Employee ID - try multiple variations
    employee_id = (
        row.get("employee_id") or
        row.get("Employee ID") or
        row.get("EmployeeID") or
        row.get("Person ID") or
        row.get("PersonID") or
        row.get("EmpID") or
        row.get("ID") or
        row.get("PersonCode") or
        row.get("Person Code") or
        ""
    ).strip()
    
    # Full Name
    full_name = (
        row.get("full_name") or
        row.get("Full Name") or
        row.get("FullName") or
        row.get("Name") or
        row.get("Person Name") or
        row.get("PersonName") or
        row.get("Employee Name") or
        ""
    ).strip()
    
    # Timestamp
    ts = (
        row.get("timestamp") or
        row.get("Time") or
        row.get("Date Time") or
        row.get("DateTime") or
        row.get("Attendance Time") or
        row.get("AttendanceTime") or
        row.get("Check Time") or
        row.get("CheckTime") or
        row.get("Event Time") or
        row.get("EventTime") or
        ""
    ).strip()
    
    # Event Type
    event_type = (
        row.get("event_type") or
        row.get("Event Type") or
        row.get("EventType") or
        row.get("Status") or
        row.get("Attendance Status") or
        row.get("AttendanceStatus") or
        row.get("Type") or
        ""
    ).strip()
    
    # Device
    device_name = (
        row.get("device_name") or
        row.get("Device") or
        row.get("Device Name") or
        row.get("DeviceName") or
        row.get("Terminal") or
        row.get("Reader") or
        device_source or
        ""
    ).strip()
    
    # Verification Mode
    verification_mode = (
        row.get("verification_mode") or
        row.get("Verification") or
        row.get("Verify Mode") or
        row.get("VerifyMode") or
        row.get("Auth Mode") or
        row.get("AuthMode") or
        row.get("Method") or
        ""
    ).strip()
    
    # Event Code
    event_code = (
        row.get("event_code") or
        row.get("Event Code") or
        row.get("EventCode") or
        row.get("Code") or
        ""
    ).strip()
    
    parsed_dt = _parse_timestamp(ts)
    
    return {
        "employee_id": employee_id,
        "full_name": full_name,
        "timestamp": parsed_dt,
        "event_type": _normalize_event_type(event_type),
        "device_name": device_name,
        "verification_mode": verification_mode,
        "event_code": event_code,
        "raw_row": row,
    }


# ===== Biometrics & Attendance Main Page =====
def admin_biometrics_attendance(request):
    """Main biometrics page: display records and import form."""
    records = AttendanceRecord.objects.all().order_by("-timestamp")[:100]
    
    kpi = {
        "present": AttendanceRecord.objects.filter(event_type=AttendanceRecord.EVENT_IN).count(),
        "late": 0,  # TODO: implement based on shift times
        "absent": 0,  # TODO: derive from employee master + scheduling
        "last_sync": AttendanceRecord.objects.order_by("-created_at").values_list("created_at", flat=True).first(),
    }
    
    context = {
        "current": "biometrics",
        "records": records,
        "kpi": kpi,
        "preview_rows": [],
        "import_errors": [],
        "import_summary": "",
    }
    
    return render(request, "admin/Biometrics_attendance.html", context)


@require_http_methods(["POST"])
def admin_biometrics_import(request):
    """
    Handle file import: validate and/or import attendance records.
    action=validate: Parse file, show preview, check for errors
    action=import: Save records to database
    """
    form = AttendanceImportForm(request.POST, request.FILES)
    
    # Get current records for display
    records = AttendanceRecord.objects.all().order_by("-timestamp")[:100]
    kpi = {
        "present": AttendanceRecord.objects.filter(event_type=AttendanceRecord.EVENT_IN).count(),
        "late": 0,
        "absent": 0,
        "last_sync": AttendanceRecord.objects.order_by("-created_at").values_list("created_at", flat=True).first(),
    }
    
    context = {
        "current": "biometrics",
        "records": records,
        "kpi": kpi,
        "preview_rows": [],
        "import_errors": [],
        "import_summary": "",
    }
    
    if not form.is_valid():
        context["import_errors"] = ["File upload failed. Please choose a valid CSV or Excel file."]
        return render(request, "admin/Biometrics_attendance.html", context)
    
    csv_file = form.cleaned_data["csv_file"]
    skip_duplicates = form.cleaned_data.get("skip_duplicates", True)
    device_source = form.cleaned_data.get("device_source", "")
    action = request.POST.get("action", "validate")
    
    # Determine file type and read
    filename = csv_file.name.lower()
    rows = []
    errors = []
    
    try:
        if filename.endswith(('.xls', '.xlsx')):
            rows = _read_excel_file(csv_file)
        else:
            rows = _read_csv_file(csv_file)
    except Exception as e:
        context["import_errors"] = [f"Error reading file: {str(e)}"]
        return render(request, "admin/Biometrics_attendance.html", context)
    
    if not rows:
        context["import_errors"] = ["File is empty or has no data rows."]
        return render(request, "admin/Biometrics_attendance.html", context)
    
    # Parse and validate all rows
    preview = []
    valid_rows = []
    validation_errors = []
    
    for i, row in enumerate(rows[:50], start=2):  # Start at row 2 (after header)
        mapped = _map_csv_row(row, device_source=device_source)
        
        # Validation
        is_valid = True
        row_errors = []
        
        if not mapped["employee_id"]:
            is_valid = False
            row_errors.append("Missing employee_id")
        
        if not mapped["timestamp"]:
            is_valid = False
            row_errors.append("Invalid/missing timestamp")
        
        status = "valid" if is_valid else "invalid"
        
        preview.append({
            "employee_id": mapped["employee_id"] or "—",
            "full_name": mapped["full_name"] or "—",
            "timestamp": mapped["timestamp"] or "—",
            "event_type": mapped["event_type"] or "—",
            "device_name": mapped["device_name"] or "—",
            "status": status,
            "errors": ", ".join(row_errors) if row_errors else "",
        })
        
        if is_valid:
            valid_rows.append(mapped)
        else:
            validation_errors.append(f"Row {i}: {', '.join(row_errors)}")
    
    context["preview_rows"] = preview
    
    # Handle validate action
    if action == "validate":
        if validation_errors:
            context["import_errors"] = validation_errors[:10]
            context["import_summary"] = f"Validation detected {len(validation_errors)} error(s). Please fix and try again."
        else:
            context["import_summary"] = f"✓ Validation passed! {len(rows)} row(s) ready to import."
        
        return render(request, "admin/Biometrics_attendance.html", context)
    
    # Handle import action
    if action != "import":
        context["import_errors"] = ["Invalid action."]
        return render(request, "admin/Biometrics_attendance.html", context)
    
    # Re-read full file for import
    csv_file.seek(0)
    try:
        if filename.endswith(('.xls', '.xlsx')):
            all_rows = _read_excel_file(csv_file)
        else:
            all_rows = _read_csv_file(csv_file)
    except Exception as e:
        context["import_errors"] = [f"Error re-reading file: {str(e)}"]
        return render(request, "admin/Biometrics_attendance.html", context)
    
    # Import all rows
    created = 0
    skipped = 0
    failed = 0
    import_errors = []
    
    for i, row in enumerate(all_rows, start=2):
        mapped = _map_csv_row(row, device_source=device_source)
        
        if not mapped["employee_id"] or not mapped["timestamp"]:
            failed += 1
            continue
        
        try:
            obj = AttendanceRecord(**mapped)
            obj.save()
            created += 1
        except IntegrityError as e:
            # Duplicate or unique constraint violation
            if skip_duplicates:
                skipped += 1
            else:
                failed += 1
                import_errors.append(f"Row {i}: Duplicate record skipped")
        except Exception as e:
            failed += 1
            import_errors.append(f"Row {i}: {str(e)}")
    
    context["import_summary"] = (
        f"✓ Import complete: {created} created | {skipped} skipped | {failed} failed"
    )
    if import_errors:
        context["import_errors"] = import_errors[:10]
    
    # Refresh records list
    context["records"] = AttendanceRecord.objects.all().order_by("-timestamp")[:100]
    
    return render(request, "admin/Biometrics_attendance.html", context)


# ===== Export endpoints =====
def admin_biometrics_template(request):
    """Download a CSV template for attendance records."""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="attendance_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        "Employee ID",
        "Full Name",
        "Date Time",
        "Event Type",
        "Device Name",
        "Verification",
        "Event Code"
    ])
    writer.writerow([
        "E001",
        "John Doe",
        "2026-02-03 08:30:00",
        "IN",
        "DS-K1T343MFWX",
        "Face",
        ""
    ])
    
    return response


def admin_biometrics_export(request):
    """Export filtered attendance records as CSV."""
    records = AttendanceRecord.objects.all().order_by("-timestamp")
    
    # Optional filtering
    employee_id = request.GET.get("employee_id", "").strip()
    event_type = request.GET.get("event_type", "").strip()
    
    if employee_id:
        records = records.filter(employee_id__icontains=employee_id)
    
    if event_type:
        records = records.filter(event_type=event_type)
    
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="attendance_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        "ID",
        "Employee ID",
        "Full Name",
        "Timestamp",
        "Event Type",
        "Device Name",
        "Verification Mode",
        "Event Code",
        "Created At"
    ])
    
    for rec in records:
        writer.writerow([
            rec.id,
            rec.employee_id,
            rec.full_name,
            rec.timestamp.isoformat() if rec.timestamp else "",
            rec.event_type,
            rec.device_name,
            rec.verification_mode,
            rec.event_code,
            rec.created_at.isoformat() if rec.created_at else "",
        ])
    
    return response


# ===== CRUD Operations =====
def attendance_list(request):
    """List all attendance records with pagination."""
    records = AttendanceRecord.objects.all().order_by("-timestamp")
    
    # Simple pagination
    page = request.GET.get("page", 1)
    per_page = 50
    start = (int(page) - 1) * per_page
    end = start + per_page
    
    total = records.count()
    paginated = records[start:end]
    
    context = {
        "current": "biometrics",
        "records": paginated,
        "page": int(page),
        "total": total,
        "per_page": per_page,
    }
    
    return render(request, "admin/attendance_list.html", context)


def attendance_detail(request, pk):
    """View a single attendance record."""
    obj = get_object_or_404(AttendanceRecord, pk=pk)
    
    context = {
        "current": "biometrics",
        "obj": obj,
    }
    
    return render(request, "admin/attendance_detail.html", context)


def attendance_create(request):
    """Create a new attendance record."""
    if request.method == "POST":
        form = AttendanceRecordForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Record created successfully!")
            return redirect("admin_biometrics")
    else:
        form = AttendanceRecordForm()
    
    context = {
        "current": "biometrics",
        "form": form,
        "mode": "create",
        "title": "Create Attendance Record"
    }
    
    return render(request, "admin/attendance_form.html", context)


def attendance_update(request, pk):
    """Edit an attendance record."""
    obj = get_object_or_404(AttendanceRecord, pk=pk)
    
    if request.method == "POST":
        form = AttendanceRecordForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Record updated successfully!")
            return redirect("admin_biometrics")
    else:
        form = AttendanceRecordForm(instance=obj)
    
    context = {
        "current": "biometrics",
        "form": form,
        "mode": "edit",
        "obj": obj,
        "title": f"Edit Record: {obj.employee_id}"
    }
    
    return render(request, "admin/attendance_form.html", context)


def attendance_delete(request, pk):
    """Delete an attendance record."""
    obj = get_object_or_404(AttendanceRecord, pk=pk)
    
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Record deleted successfully!")
        return redirect("admin_biometrics")
    
    context = {
        "current": "biometrics",
        "obj": obj,
    }
    
    return render(request, "admin/attendance_delete.html", context)
